"""Parse the recipe workbook (one sheet per recipe) or a flat recipe CSV.

Workbook layout (per sheet):
    row  1: recipe name
    row  3: "Yield:"  <qty> ... <unit>
    row  4: "Portion:" <qty> ... <unit> ... "# Portions:" <n>
    ...
    header row: Ingredient | Qty | Measure | Yld % | Instructions | Cost | Cost Per
    ingredient rows follow until a row that has cost but no ingredient (total row)

Sheet-name prefixes: ``MI`` menu item, ``PREP``/``BATCH``/``YIELD`` sub-recipes.
"""
from __future__ import annotations

import warnings
from pathlib import Path
from typing import Any

import pandas as pd

from foodcost.utils.text import strip_recipe_category

warnings.filterwarnings("ignore", module="openpyxl")

RECIPE_TYPES = {"MI": "menu_item", "PREP": "prep", "BATCH": "batch", "YIELD": "yield"}


def _sheet_type(sheet_name: str) -> str:
    prefix = sheet_name.split()[0] if sheet_name.split() else ""
    return RECIPE_TYPES.get(prefix, "menu_item")


def _find_value(row: tuple, start: int = 0) -> Any:
    """First non-empty value in a row after index ``start``."""
    for j in range(start, len(row)):
        if row[j] not in (None, ""):
            return row[j]
    return None


def parse_recipe_workbook(path: str | Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Parse a multi-sheet recipe workbook.

    Returns
    -------
    recipes : DataFrame [recipe, recipe_type, yield_qty, yield_unit,
                         portion_qty, portion_unit, n_portions, book_cost]
    ingredients : DataFrame [recipe, ingredient, category, qty, measure,
                             yield_pct, book_cost, book_cost_per_portion]
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    recipe_rows: list[dict] = []
    ing_rows: list[dict] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        meta: dict[str, Any] = {
            "recipe": sheet.strip(),
            "recipe_type": _sheet_type(sheet),
            "yield_qty": None, "yield_unit": None,
            "portion_qty": None, "portion_unit": None, "n_portions": None,
        }
        header_idx = None
        col: dict[str, int] = {}

        for i, row in enumerate(rows):
            if not row:
                continue
            cells = {j: v for j, v in enumerate(row) if v not in (None, "")}
            texts = {j: str(v).strip() for j, v in cells.items() if isinstance(v, str)}

            if "Yield:" in texts.values():
                nums = [v for v in cells.values() if isinstance(v, (int, float))]
                units = [t for t in texts.values() if t not in ("Yield:",)]
                meta["yield_qty"] = nums[0] if nums else None
                meta["yield_unit"] = units[0] if units else None
            elif any(t.startswith("Portion:") for t in texts.values()):
                idxs = sorted(cells)
                nums = [(j, cells[j]) for j in idxs if isinstance(cells[j], (int, float))]
                strs = [(j, texts[j]) for j in texts
                        if texts[j] not in ("Portion:", "# Portions:")]
                if nums:
                    meta["portion_qty"] = nums[0][1]
                    if len(nums) > 1:
                        meta["n_portions"] = nums[-1][1]
                if strs:
                    meta["portion_unit"] = strs[0][1]
            elif texts.get(0) == "Ingredient":
                header_idx = i
                for j, t in texts.items():
                    col[t] = j
                break

        book_total = None
        if header_idx is not None:
            qty_col = col.get("Qty", 12)
            measure_col = col.get("Measure", 21)
            yld_col = col.get("Yld %", 27)
            cost_col = col.get("Cost", 41)
            cost_per_col = col.get("Cost Per", 43)

            for row in rows[header_idx + 1:]:
                if not row:
                    break
                name = row[0] if len(row) > 0 else None
                cost = row[cost_col] if len(row) > cost_col else None
                if (name in (None, "")) and isinstance(cost, (int, float)):
                    book_total = float(cost)          # total row
                    break
                if name in (None, ""):
                    # blank spacer without total -> keep scanning one more row
                    continue
                if str(name).strip() in ("Additional Info",):
                    break
                cat, clean = strip_recipe_category(str(name).strip())
                ing_rows.append({
                    "recipe": meta["recipe"],
                    "ingredient": str(name).strip(),
                    "ingredient_clean": clean,
                    "category": cat,
                    "qty": row[qty_col] if len(row) > qty_col else None,
                    "measure": row[measure_col] if len(row) > measure_col else None,
                    "yield_pct": row[yld_col] if len(row) > yld_col else None,
                    "book_cost": cost if isinstance(cost, (int, float)) else None,
                    "book_cost_per_portion": (
                        row[cost_per_col]
                        if len(row) > cost_per_col and isinstance(row[cost_per_col], (int, float))
                        else None
                    ),
                })

        meta["book_cost"] = book_total
        recipe_rows.append(meta)

    wb.close()
    recipes = pd.DataFrame(recipe_rows)
    ingredients = pd.DataFrame(ing_rows)
    return recipes, ingredients


def parse_recipe_csv(path: str | Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Parse a flat CSV/Excel recipe file (sample-data format).

    Expected columns: recipe, recipe_type, ingredient, category, qty, measure,
    optional: yield_qty, yield_unit, book_cost
    """
    p = Path(path)
    df = pd.read_excel(p) if p.suffix.lower() in (".xlsx", ".xls") else pd.read_csv(p)
    df.columns = [c.strip().lower() for c in df.columns]
    required = {"recipe", "ingredient", "qty", "measure"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Recipe file missing columns: {sorted(missing)}")

    if "category" not in df.columns:
        df["category"] = [strip_recipe_category(x)[0] for x in df["ingredient"]]
    if "recipe_type" not in df.columns:
        df["recipe_type"] = "menu_item"
    df["ingredient_clean"] = [strip_recipe_category(str(x))[1] for x in df["ingredient"]]
    if "book_cost" not in df.columns:
        df["book_cost"] = None
    df["yield_pct"] = df.get("yield_pct", 100.0)
    df["book_cost_per_portion"] = df["book_cost"]

    grp = df.groupby(["recipe", "recipe_type"], as_index=False).agg(
        book_cost=("book_cost", "sum"))
    grp["yield_qty"] = df.groupby(["recipe", "recipe_type"])["qty"].transform(
        lambda s: 1.0).iloc[:len(grp)].values if len(grp) else []
    recipes = grp.assign(yield_qty=1.0, yield_unit="Each",
                         portion_qty=1.0, portion_unit="Each", n_portions=1.0)
    ingredients = df[["recipe", "ingredient", "ingredient_clean", "category",
                      "qty", "measure", "yield_pct", "book_cost",
                      "book_cost_per_portion"]].copy()
    return recipes, ingredients
