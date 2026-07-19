"""POS ingestion: sales-mix exports, item-sales-detail exports, generic CSV.

Supported inputs
----------------
1. ``slsmix*.xls``  — per-location Sales Mix Report (Description, ID, Qty,
   Sales $ ...), with the location in the "Criteria:" header row.
2. ``Item Sales Detail.xlsx`` — transaction-level export (Business Date,
   Order Id, Sold Name, Sold Price) used for daily trend analysis.
3. Generic CSV/Excel with columns: item, qty_sold, revenue and optional
   location/date/category.
"""
from __future__ import annotations

import re
import warnings
from pathlib import Path

import pandas as pd

from foodcost.config import normalize_location

warnings.filterwarnings("ignore")

_SUMMARY_ROWS = {
    "FILTERED COUNT", "GRAND TOTAL COUNT", "HIGHEST", "AVERAGE", "LOWEST",
}


def parse_sales_mix(path: str | Path) -> pd.DataFrame:
    """Parse one slsmix export -> [location, item, pos_id, qty_sold, revenue]."""
    raw = pd.read_excel(str(path), header=None)

    # Location lives in the "Criteria:" row; period in the row above it.
    location, period = "Unknown", None
    for i in range(min(6, len(raw))):
        cell = str(raw.iloc[i, 0])
        if cell.startswith("Criteria:"):
            location = normalize_location(
                cell.replace("Criteria:", "").replace("\xa0", "").strip(" ;"))
        m = re.match(r"(\d{2}/\d{2}/\d{4}) - (\d{2}/\d{2}/\d{4})", cell)
        if m:
            period = (pd.Timestamp(m.group(1)), pd.Timestamp(m.group(2)))

    # Find the column-header row ("Description" in col 0).
    hdr = None
    for i in range(len(raw)):
        if str(raw.iloc[i, 0]).strip() == "Description":
            hdr = i
            break
    if hdr is None:
        return pd.DataFrame()

    body = raw.iloc[hdr + 1:].copy()
    body.columns = range(body.shape[1])
    rows = []
    for _, r in body.iterrows():
        desc = r[0]
        if not isinstance(desc, str) or desc.strip().upper() in _SUMMARY_ROWS:
            continue
        try:
            qty = float(r[2])
            revenue = float(r[4]) if pd.notna(r[4]) else 0.0
        except (TypeError, ValueError):
            continue
        if qty == 0 and revenue == 0:
            continue
        rows.append({
            "location": location,
            "item": desc.strip(),
            "pos_id": str(r[1]) if pd.notna(r[1]) else None,
            "qty_sold": qty,
            "revenue": revenue,
            "period_start": period[0] if period else None,
            "period_end": period[1] if period else None,
        })
    return pd.DataFrame(rows)


def parse_sales_mix_folder(folder: str | Path) -> pd.DataFrame:
    """Parse every slsmix*.xls file in a folder."""
    frames = []
    for f in sorted(Path(folder).glob("slsmix*.xls*")):
        try:
            df = parse_sales_mix(f)
            if not df.empty:
                frames.append(df)
        except Exception:
            continue
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def parse_item_sales_detail(path: str | Path) -> pd.DataFrame:
    """Parse the transaction-level Item Sales Detail export.

    Returns [date, item, orders, qty_sold, revenue] aggregated by day+item.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    recs = []
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 18:
            continue
        date, name, price = row[1], row[9], row[17]
        if not hasattr(date, "year") or not isinstance(name, str):
            continue
        recs.append((pd.Timestamp(date), name.strip(),
                     float(price) if isinstance(price, (int, float)) else 0.0))
    wb.close()
    if not recs:
        return pd.DataFrame()
    df = pd.DataFrame(recs, columns=["date", "item", "revenue"])
    out = (df.groupby(["date", "item"], as_index=False)
             .agg(qty_sold=("revenue", "size"), revenue=("revenue", "sum")))
    return out


def parse_generic_pos(path: str | Path) -> pd.DataFrame:
    """Parse a generic CSV/Excel POS export with flexible column names."""
    p = Path(path)
    df = pd.read_excel(p) if p.suffix.lower() in (".xlsx", ".xls") else pd.read_csv(p)
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

    aliases = {
        "item": ["item", "menu_item", "description", "sold_name", "name", "product"],
        "qty_sold": ["qty_sold", "qty", "quantity", "quantity_sold", "count"],
        "revenue": ["revenue", "sales", "amount", "net_sales", "sales_$"],
        "location": ["location", "store", "site", "unit"],
        "date": ["date", "business_date", "day"],
    }
    resolved: dict[str, str] = {}
    for canon, options in aliases.items():
        for opt in options:
            if opt in df.columns:
                resolved[canon] = opt
                break
    if "item" not in resolved or "qty_sold" not in resolved:
        raise ValueError("POS file needs at least an item column and a quantity column")

    out = pd.DataFrame({
        "item": df[resolved["item"]].astype(str).str.strip(),
        "qty_sold": pd.to_numeric(df[resolved["qty_sold"]], errors="coerce").fillna(0),
        "revenue": pd.to_numeric(df[resolved["revenue"]], errors="coerce").fillna(0)
        if "revenue" in resolved else 0.0,
        "location": df[resolved["location"]].astype(str) if "location" in resolved else "All",
    })
    if "date" in resolved:
        out["date"] = pd.to_datetime(df[resolved["date"]], errors="coerce")
    return out
